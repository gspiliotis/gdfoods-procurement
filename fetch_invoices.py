#!/usr/bin/env python3
"""
Fetch invoice data from myDATA API and generate Excel spreadsheet with aggregated quantities.
"""
import argparse
import os
import sys
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional
import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# API Constants - load from environment variables
USER_ID = os.getenv("MYDATA_USER_ID")
API_KEY = os.getenv("MYDATA_API_KEY")
API_BASE_URL = "https://mydatapi.aade.gr/myDATA/RequestDocs"

if not USER_ID or not API_KEY:
    print("Error: MYDATA_USER_ID and MYDATA_API_KEY environment variables must be set", file=sys.stderr)
    print("Please create a .env file with your credentials (see .env.example)", file=sys.stderr)
    sys.exit(1)


def convert_date_to_api_format(date_str: str) -> str:
    """
    Convert date from YYYY-MM-DD to DD/MM/YYYY format for API.

    Args:
        date_str: Date in YYYY-MM-DD format

    Returns:
        Date in DD/MM/YYYY format
    """
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    return date_obj.strftime("%d/%m/%Y")


def fetch_invoices(
    date_from: str,
    date_to: str,
    receiver_vat_number: Optional[str] = None,
    next_partition_key: Optional[str] = None,
    next_row_key: Optional[str] = None
) -> str:
    """
    Fetch invoices from myDATA API.

    Args:
        date_from: Start date in YYYY-MM-DD format
        date_to: End date in YYYY-MM-DD format
        receiver_vat_number: VAT number of the receiver (optional, fetches all if None)
        next_partition_key: Pagination key for next partition
        next_row_key: Pagination key for next row

    Returns:
        XML response as string
    """
    # Convert dates to DD/MM/YYYY format for API
    api_date_from = convert_date_to_api_format(date_from)
    api_date_to = convert_date_to_api_format(date_to)

    params = {
        "mark": "1",
        "dateFrom": api_date_from,
        "dateTo": api_date_to,
    }

    if receiver_vat_number:
        params["receiverVatNumber"] = receiver_vat_number

    if next_partition_key:
        params["nextPartitionKey"] = next_partition_key
    if next_row_key:
        params["nextRowKey"] = next_row_key

    headers = {
        "aade-user-id": USER_ID,
        "Ocp-Apim-Subscription-Key": API_KEY
    }

    try:
        response = requests.get(API_BASE_URL, params=params, headers=headers, timeout=30)
        response.raise_for_status()
        return response.text
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data for VAT {receiver_vat_number}: {e}", file=sys.stderr)
        if hasattr(e, 'response') and e.response is not None:
            print(f"Response text: {e.response.text}", file=sys.stderr)
        return ""


def parse_invoices(xml_content: str) -> Tuple[List[Dict], Optional[str], Optional[str]]:
    """
    Parse XML response and extract invoice data.

    Args:
        xml_content: XML response as string

    Returns:
        Tuple of (list of invoice records, next_partition_key, next_row_key)
    """
    if not xml_content:
        return [], None, None

    try:
        root = ET.fromstring(xml_content)
    except ET.ParseError as e:
        print(f"Error parsing XML: {e}", file=sys.stderr)
        return [], None, None

    # Define namespace
    ns = {'ns': 'http://www.aade.gr/myDATA/invoice/v1.0'}

    # Extract pagination tokens
    next_partition_key = None
    next_row_key = None

    continuation_token = root.find("ns:continuationToken", ns)
    if continuation_token is not None:
        npk = continuation_token.find("ns:nextPartitionKey", ns)
        nrk = continuation_token.find("ns:nextRowKey", ns)
        if npk is not None:
            next_partition_key = npk.text
        if nrk is not None:
            next_row_key = nrk.text

    # Extract invoice data
    records = []
    # Find invoicesDoc container first
    invoices_doc = root.find("ns:invoicesDoc", ns)
    if invoices_doc is None:
        return records, next_partition_key, next_row_key

    invoices = invoices_doc.findall("ns:invoice", ns)

    for invoice in invoices:
        # Get issuer name
        issuer = invoice.find("ns:issuer", ns)
        if issuer is None:
            continue

        issuer_vat_elem = issuer.find("ns:vatNumber", ns)
        issuer_vat = issuer_vat_elem.text.strip() if issuer_vat_elem is not None and issuer_vat_elem.text else ""

        issuer_name_elem = issuer.find("ns:name", ns)
        if issuer_name_elem is None or not issuer_name_elem.text:
            continue
        issuer_name = issuer_name_elem.text.strip()

        # Get issue date
        invoice_header = invoice.find("ns:invoiceHeader", ns)
        if invoice_header is None:
            continue

        issue_date_elem = invoice_header.find("ns:issueDate", ns)
        if issue_date_elem is None or not issue_date_elem.text:
            continue
        issue_date = issue_date_elem.text.strip()

        # Get invoice details
        for detail in invoice.findall("ns:invoiceDetails", ns):
            item_descr_elem = detail.find("ns:itemDescr", ns)
            quantity_elem = detail.find("ns:quantity", ns)

            if item_descr_elem is None or quantity_elem is None:
                continue

            if not item_descr_elem.text or not quantity_elem.text:
                continue

            item_descr = item_descr_elem.text.strip()
            try:
                quantity = float(quantity_elem.text)
            except ValueError:
                continue

            records.append({
                "issuer_name": issuer_name,
                "issuer_vat": issuer_vat,
                "item_descr": item_descr,
                "issue_date": issue_date,
                "quantity": quantity
            })

    return records, next_partition_key, next_row_key


def fetch_all_invoices_for_period(date_from: str, date_to: str) -> List[Dict]:
    """
    Fetch all invoices for the given date range (no VAT filtering).

    Args:
        date_from: Start date in YYYY-MM-DD format
        date_to: End date in YYYY-MM-DD format

    Returns:
        List of all invoice records
    """
    all_records = []
    next_partition_key = None
    next_row_key = None
    page = 1

    print(f"Fetching all invoices for period {date_from} to {date_to}...")

    while True:
        xml_content = fetch_invoices(
            date_from, date_to,
            next_partition_key=next_partition_key,
            next_row_key=next_row_key
        )

        if not xml_content:
            break

        records, next_partition_key, next_row_key = parse_invoices(xml_content)
        all_records.extend(records)

        print(f"  Page {page}: Found {len(records)} invoice items")
        page += 1

        if not next_partition_key or not next_row_key:
            break

    print(f"Total invoice items fetched: {len(all_records)}")
    return all_records


def filter_by_vat_numbers(records: List[Dict], vat_data: List[Tuple[str, int]]) -> List[Dict]:
    """
    Filter records by issuer VAT numbers and apply date adjustments.

    Args:
        records: List of invoice records (must contain 'issuer_vat' field)
        vat_data: List of tuples (vat_number, date_adjustment)

    Returns:
        Filtered list of records with date_adjustment set
    """
    vat_map = {vat.strip(): adj for vat, adj in vat_data}
    filtered = []
    for record in records:
        issuer_vat = record.get("issuer_vat", "")
        if issuer_vat in vat_map:
            record["date_adjustment"] = vat_map[issuer_vat]
            filtered.append(record)
    return filtered


def get_greek_day_name(date_str: str) -> str:
    """
    Get Greek day name from date string.

    Args:
        date_str: Date in YYYY-MM-DD format

    Returns:
        Greek day name
    """
    greek_days = {
        0: "Δευτέρα",     # Monday
        1: "Τρίτη",       # Tuesday
        2: "Τετάρτη",     # Wednesday
        3: "Πέμπτη",      # Thursday
        4: "Παρασκευή",   # Friday
        5: "Σάββατο",     # Saturday
        6: "Κυριακή"      # Sunday
    }
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    return greek_days[date_obj.weekday()]


def aggregate_data(records: List[Dict]) -> Tuple[Dict[Tuple[str, str], Dict[str, float]], List[str]]:
    """
    Aggregate quantities by (issuer, item, date) with date adjustment applied.

    Args:
        records: List of invoice records with date_adjustment field

    Returns:
        Tuple of (aggregated data dict, sorted list of dates)
    """
    # Dictionary: (issuer_name, item_descr) -> {date: quantity}
    aggregated = defaultdict(lambda: defaultdict(float))
    dates_set = set()

    for record in records:
        key = (record["issuer_name"], record["item_descr"])
        original_date = record["issue_date"]
        quantity = record["quantity"]
        date_adjustment = record.get("date_adjustment", 0)

        # Apply date adjustment
        if date_adjustment != 0:
            date_obj = datetime.strptime(original_date, "%Y-%m-%d")
            adjusted_date_obj = date_obj + timedelta(days=date_adjustment)
            adjusted_date = adjusted_date_obj.strftime("%Y-%m-%d")
        else:
            adjusted_date = original_date

        aggregated[key][adjusted_date] += quantity
        dates_set.add(adjusted_date)

    # Sort dates
    sorted_dates = sorted(list(dates_set))

    return aggregated, sorted_dates


def generate_excel(aggregated_data: Dict, dates: List[str], output_file: str):
    """
    Generate Excel file with aggregated data.

    Args:
        aggregated_data: Dictionary of aggregated quantities
        dates: Sorted list of dates
        output_file: Output Excel file path
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Quantities"

    # Write first header row (dates)
    ws.cell(1, 1, "Issuer")
    ws.cell(1, 2, "Item Description")
    for col_idx, date in enumerate(dates, start=3):
        ws.cell(1, col_idx, date)

    # Write second header row (day names in Greek)
    for col_idx, date in enumerate(dates, start=3):
        ws.cell(2, col_idx, get_greek_day_name(date))

    # Make headers bold
    for col in range(1, len(dates) + 3):
        ws.cell(1, col).font = Font(bold=True)
        ws.cell(2, col).font = Font(bold=True)

    # Write data rows
    row_idx = 3
    for (issuer_name, item_descr), date_quantities in sorted(aggregated_data.items()):
        ws.cell(row_idx, 1, issuer_name)
        ws.cell(row_idx, 2, item_descr)

        for col_idx, date in enumerate(dates, start=3):
            quantity = date_quantities.get(date, 0)
            if quantity > 0:
                ws.cell(row_idx, col_idx, quantity)

        row_idx += 1

    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 60
    for col_idx in range(3, len(dates) + 3):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    wb.save(output_file)
    print(f"\nExcel file generated: {output_file}")


def generate_csv(aggregated_data: Dict, dates: List[str], output_file: str):
    """
    Generate CSV file with aggregated data.

    Args:
        aggregated_data: Dictionary of aggregated quantities
        dates: Sorted list of dates
        output_file: Output CSV file path
    """
    with open(output_file, 'w', encoding='utf-8') as f:
        # Write first header row (dates)
        f.write(";;")
        f.write(";".join(dates))
        f.write("\n")

        # Write second header row (day names in Greek)
        f.write(";;")
        f.write(";".join([get_greek_day_name(date) for date in dates]))
        f.write("\n")

        # Write data rows
        for (issuer_name, item_descr), date_quantities in sorted(aggregated_data.items()):
            row = [issuer_name, item_descr]
            for date in dates:
                quantity = date_quantities.get(date, 0)
                row.append(str(int(quantity)) if quantity > 0 else "")

            f.write(";".join(row))
            f.write("\n")

    print(f"\nCSV file generated: {output_file}")


def read_vat_numbers(filename: str) -> List[Tuple[str, int]]:
    """
    Read VAT numbers and date adjustments from file.
    Lines starting with # are treated as comments and ignored.
    Anything after a # character on a line is also ignored.

    Format: VAT_NUMBER DATE_ADJUSTMENT
    Example: 094254743  -1

    Args:
        filename: Path to file containing VAT numbers and date adjustments

    Returns:
        List of tuples (vat_number, date_adjustment)
    """
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            vat_data = []
            line_num = 0
            for line in f:
                line_num += 1
                # Remove comments (anything after #)
                line = line.split('#')[0].strip()
                # Skip empty lines
                if not line:
                    continue

                # Parse VAT number and date adjustment
                parts = line.split()
                if len(parts) < 2:
                    print(f"Warning: Line {line_num} missing date adjustment, using 0: {line}", file=sys.stderr)
                    vat_number = parts[0]
                    date_adjustment = 0
                else:
                    vat_number = parts[0]
                    try:
                        date_adjustment = int(parts[1])
                    except ValueError:
                        print(f"Warning: Line {line_num} has invalid date adjustment, using 0: {line}", file=sys.stderr)
                        date_adjustment = 0

                vat_data.append((vat_number, date_adjustment))
            return vat_data
    except FileNotFoundError:
        print(f"Error: File '{filename}' not found", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Error reading file '{filename}': {e}", file=sys.stderr)
        sys.exit(1)


def validate_date(date_str: str) -> bool:
    """Validate date format (YYYY-MM-DD)."""
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
        return True
    except ValueError:
        return False


def write_vat_output(records: List[Dict], output_file: str):
    """
    Write unique issuer VAT numbers from records to a file.

    Args:
        records: List of invoice records
        output_file: Path to the output file
    """
    # Collect unique VATs with their issuer names
    vat_info = {}
    for record in records:
        vat = record.get("issuer_vat", "")
        if vat and vat not in vat_info:
            vat_info[vat] = record.get("issuer_name", "")

    with open(output_file, 'w', encoding='utf-8') as f:
        for vat in sorted(vat_info.keys()):
            name = vat_info[vat]
            f.write(f"{vat}  0   # {name}\n")

    print(f"\nVAT numbers file generated: {output_file} ({len(vat_info)} unique VAT numbers)")


def main():
    parser = argparse.ArgumentParser(
        description="Fetch invoice data from myDATA API and generate Excel/CSV report"
    )
    parser.add_argument(
        "start_date",
        help="Start date in YYYY-MM-DD format"
    )
    parser.add_argument(
        "end_date",
        help="End date in YYYY-MM-DD format"
    )
    parser.add_argument(
        "vat_file",
        nargs="?",
        default=None,
        help="File containing VAT numbers to filter by (optional; if omitted, all invoices are fetched)"
    )
    parser.add_argument(
        "-o", "--output",
        default="invoice_report",
        help="Output file name (without extension, default: invoice_report)"
    )
    parser.add_argument(
        "-f", "--format",
        choices=["xlsx", "csv", "both"],
        default="xlsx",
        help="Output format (default: xlsx)"
    )
    parser.add_argument(
        "--vat-out",
        default=None,
        help="Output file to write unique VAT numbers found in fetched invoices (optional)"
    )

    args = parser.parse_args()

    # Validate dates
    if not validate_date(args.start_date):
        print(f"Error: Invalid start date '{args.start_date}'. Use YYYY-MM-DD format.", file=sys.stderr)
        sys.exit(1)

    if not validate_date(args.end_date):
        print(f"Error: Invalid end date '{args.end_date}'. Use YYYY-MM-DD format.", file=sys.stderr)
        sys.exit(1)

    print(f"Date range: {args.start_date} to {args.end_date}\n")

    # Fetch all invoices for the period
    all_records = fetch_all_invoices_for_period(args.start_date, args.end_date)

    if not all_records:
        print("\nNo invoice data found")
        sys.exit(0)

    # Write VAT output file if requested
    if args.vat_out:
        write_vat_output(all_records, args.vat_out)

    # Filter by VAT numbers if file provided
    if args.vat_file:
        vat_data = read_vat_numbers(args.vat_file)
        if not vat_data:
            print("Error: No VAT numbers found in file", file=sys.stderr)
            sys.exit(1)
        print(f"\nFiltering by {len(vat_data)} VAT number(s) from {args.vat_file}")
        records = filter_by_vat_numbers(all_records, vat_data)
        print(f"Records after filtering: {len(records)}")
    else:
        # No filtering - set date_adjustment to 0 for all records
        records = all_records
        for r in records:
            r["date_adjustment"] = 0

    if not records:
        print("\nNo invoice data after filtering")
        sys.exit(0)

    # Aggregate data
    aggregated_data, dates = aggregate_data(records)
    print(f"\nUnique (issuer, item) combinations: {len(aggregated_data)}")
    print(f"Date range in data: {dates[0]} to {dates[-1]}" if dates else "No dates")

    # Generate output
    if args.format in ["xlsx", "both"]:
        generate_excel(aggregated_data, dates, f"{args.output}.xlsx")

    if args.format in ["csv", "both"]:
        generate_csv(aggregated_data, dates, f"{args.output}.csv")

    print("\nDone!")


if __name__ == "__main__":
    main()
